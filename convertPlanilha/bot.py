# Import for the Web Bot
from botcity.web import WebBot, Browser, By

# Import for integration with BotCity Maestro SDK
from botcity.maestro import *

# Disable errors if we are not connected to Maestro
BotMaestroSDK.RAISE_NOT_CONNECTED = False

# import modules
from metodos.function_planilha import *
from metodos.get_convert_bcb import *
from openpyxl.styles import Border, Side

new_caminho = 'teste.xlsx'


def main():
    # Runner passes the server url, the id of the task being executed,
    # the access token and the parameters that this task receives (when applicable).
    maestro = BotMaestroSDK.from_sys_args()
    ## Fetch the BotExecution with details from the task, including parameters
    execution = maestro.get_execution()

    print(f"Task ID is: {execution.task_id}")
    print(f"Task Parameters are: {execution.parameters}")

    bot = WebBot()
    

    # Configure whether or not to run on headless mode
    bot.headless = False

    # Uncomment to change the default Browser to Firefox
    bot.browser = Browser.FIREFOX

    # Uncomment to set the WebDriver path
    bot.driver_path = r"/home/lukeid/Documentos/geckodriver"
    


    # Opens the BotCity website.
    bot.browse("https://www.botcity.dev")

    # Implement here your logic...
    
    nova_aba = op.Workbook().active
    nova_aba.title = 'Valores Convertidos'

    taxas_conversao = obter_taxas_de_conversao()

    planilhas = [
        {'nome': 'Copper', 'key': 'Average'},
        {'nome': 'Aluminium Alloy', 'key': 'Average'},
        {'nome': 'Primary Aluminium', 'key': 'Average'}
    ]

    for planilha_info in planilhas:
        nova_aba = input_values(planilha_info['nome'], planilha_info['key'], taxas_conversao, nova_aba)
        nova_aba.append([])  # Adicionar uma linha em branco para separar as tabelas

    # Ajustar largura das colunas
    for col in nova_aba.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        nova_aba.column_dimensions[column].width = adjusted_width

    border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

    # Aplicar borda às células
    for row in nova_aba.iter_rows(min_row=1, max_row=nova_aba.max_row, min_col=1, max_col=nova_aba.max_column):
        for cell in row:
            cell.border = border_style

    # Salvar o arquivo
    nova_aba.parent.save('valores_convertidos_com_bordas.xlsx')


        

    # Wait 3 seconds before closing

    # Finish and clean up the Web Browser
    # You MUST invoke the stop_browser to avoid
    # leaving instances of the webdriver open
    bot.stop_browser()

    # Uncomment to mark this task as finished on BotMaestro
    # maestro.finish_task(
    #     task_id=execution.task_id,
    #     status=AutomationTaskFinishStatus.SUCCESS,
    #     message="Task Finished OK."
    # )


def not_found(label):
    print(f"Element not found: {label}")


if __name__ == '__main__':
    main()
